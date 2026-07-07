import os
import tempfile
from pathlib import Path

TEXT_PDF_MIME = "application/pdf"
TEXT_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
TEXT_DOC_MIME = "application/msword"
TEXT_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TEXT_XLS_MIME = "application/vnd.ms-excel"

SCANNED_PDF_MIN_CHARS = 20


def extract_text(file_path: str, mime_type: str) -> str | None:
    if mime_type == TEXT_PDF_MIME:
        return _extract_pdf(file_path)
    if mime_type in (TEXT_DOCX_MIME, TEXT_DOC_MIME):
        return _extract_docx(file_path)
    if mime_type in (TEXT_XLSX_MIME, TEXT_XLS_MIME):
        return _extract_xlsx(file_path)
    if mime_type and mime_type.startswith("image/"):
        return _extract_image(file_path)
    return None


def _extract_pdf(file_path: str) -> str | None:
    try:
        import fitz
    except ImportError:
        return None

    try:
        doc = fitz.open(file_path)
    except Exception:
        return None

    text_parts = []
    for page in doc:
        text_parts.append(page.get_text())

    doc.close()
    plain = "\n".join(text_parts).strip()

    if len(plain) < SCANNED_PDF_MIN_CHARS:
        ocr_text = _ocr_pdf(file_path)
        if ocr_text:
            return ocr_text

    return plain if plain else None


def _ocr_pdf(file_path: str) -> str | None:
    try:
        import fitz
        import easyocr
    except ImportError:
        return None

    try:
        reader = easyocr.Reader(["en"], gpu=False)
        doc = fitz.open(file_path)
    except Exception:
        return None

    text_parts = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        img_bytes = pix.tobytes("png")
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            tmp.write(img_bytes)
            tmp_path = tmp.name
        try:
            result = reader.readtext(tmp_path, detail=0, paragraph=True)
            text_parts.extend(result)
        finally:
            os.unlink(tmp_path)

    doc.close()
    combined = "\n".join(text_parts).strip()
    return combined if combined else None


def _extract_docx(file_path: str) -> str | None:
    try:
        from docx import Document
    except ImportError:
        return None

    try:
        doc = Document(file_path)
    except Exception:
        return None

    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n".join(paragraphs).strip()
    return text if text else None


def _extract_xlsx(file_path: str) -> str | None:
    try:
        import openpyxl
    except ImportError:
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return None

    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            joined = "\t".join(cells).strip()
            if joined:
                rows.append(joined)
        if rows:
            lines.append(f"=== {sheet_name} ===")
            lines.extend(rows)

    wb.close()
    text = "\n".join(lines).strip()
    return text if text else None


def _extract_image(file_path: str) -> str | None:
    try:
        import easyocr
    except ImportError:
        return None

    try:
        reader = easyocr.Reader(["en"], gpu=False)
        result = reader.readtext(file_path, detail=0, paragraph=True)
    except Exception:
        return None

    text = "\n".join(result).strip()
    return text if text else None
